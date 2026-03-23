from telegram import Update
from telegram.ext import ConversationHandler, MessageHandler, CommandHandler, ContextTypes, filters
from datetime import datetime
from openpyxl import Workbook
from openpyxl import Font, PatternFill, Alignment, Border, Side
import os

from database.db import user_exists, get_overtime_by_moth, get_user_by_telegram_id

REGISTER_MES = 1

# Columnas de la planilla oficial
HEADERS = [
    "LEGAJO",
    "APELLIDO Y NOMBRE",
    "SECTOR",
    "Dia y horario habitual",
    "FECHA",
    "Dia",
    "HORARIO de inicio de hora extra",
    "HORARIO de finaliazcion de hora extra",
    "DESCRIPCIÓN DE TAREAS",
    "ISSUE/TICKET ASOCIADO",
    "PROYECTO O SERVICIO?",
    "Nombre del proyecto",
    "Cliente",
    "QUIEN SOLICITÓ",
    "QUIEN VALIDÓ",
    "Dias de viaje",
    "Feriados  trabajados (Jornada completa)",
    "DÍAS GUARDIAS",
    "HORAS FERIADOS (Jornadas parciales)",
    "HORAS EXTRAS 50%",
    "HORAS EXTRAS 100%",
    "Notas u observaciones. ",
    "Revision del lider",
]
 
DIAS_ES = {
    0: "lunes", 1: "martes", 2: "miércoles",
    3: "jueves", 4: "viernes", 5: "sábado", 6: "domingo"
}
 
def _formato_hora(hora_str: str) -> str:
    """Convierte HH:MM a formato 12h con a.m./p.m. como en la planilla."""
    try:
        h = datetime.strptime(hora_str, "%H:%M")
        hora12 = h.strftime("%I:%M %p").lower().replace("am", "a.m.").replace("pm", "p.m.")
        # Quitar cero inicial: "06:00" -> "6:00"
        if hora12.startswith("0"):
            hora12 = hora12[1:]
        return hora12
    except ValueError:
        return hora_str
 
def _formato_fecha(fecha_str: str) -> str:
    """Convierte DD/MM/YYYY a M/D/YYYY como en la planilla."""
    try:
        d = datetime.strptime(fecha_str, "%d/%m/%Y")
        return f"{d.month}/{d.day}/{d.year}"
    except ValueError:
        return fecha_str
 
def _dia_semana(fecha_str: str) -> str:
    """Devuelve el nombre del día en español."""
    try:
        d = datetime.strptime(fecha_str, "%d/%m/%Y")
        return DIAS_ES[d.weekday()]
    except ValueError:
        return ""
 
def _build_xlsx(user_data: dict, registros: list, month: int, year: int) -> str:
    """Genera el xlsx en formato de la planilla oficial. Devuelve el filepath."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Horas Extras"
 
    # --- Estilos ---
    header_font = Font(name="Arial", bold=True, size=9)
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font_white = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    # Fila vacía inicial (como en el original)
    ws.append([""] * len(HEADERS))
 
    # --- Encabezados ---
    ws.append(HEADERS)
    header_row = ws.max_row
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
 
    # --- Filas de datos ---
    legajo   = user_data.get("legajo", "")
    nombre   = user_data.get("nombre", "")
    sector   = user_data.get("area", "")
    jornada  = user_data.get("jornada", "")
 
    for r in registros:
        # r = (fecha, hora_inicio, hora_fin, descripcion, ticket, cliente)
        fecha, hora_i, hora_f, descripcion, ticket, cliente, proyecto = r
 
        row = [
            legajo,                          # LEGAJO
            nombre,                          # APELLIDO Y NOMBRE
            sector,                          # SECTOR
            jornada,                         # Dia y horario habitual
            _formato_fecha(fecha),           # FECHA
            _dia_semana(fecha),              # Dia
            _formato_hora(hora_i),           # HORARIO inicio
            _formato_hora(hora_f),           # HORARIO fin
            descripcion,                     # DESCRIPCIÓN
            ticket,                          # TICKET
            "SERVICIO",                      # PROYECTO O SERVICIO (valor por defecto)
            proyecto,                        # Nombre del proyecto
            cliente,                         # Cliente
            "", "", "", "", "", "", "", "", "", ""  # Resto en blanco
        ]
        ws.append(row)
        data_row = ws.max_row
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
 
    # --- Fila TOTALES ---
    ws.append([""] * len(HEADERS))
    totales_row = ws.max_row
    ws.cell(row=totales_row, column=14).value = "TOTALES"
    ws.cell(row=totales_row, column=14).font = Font(name="Arial", bold=True, size=9)
 
    # --- Anchos de columna ---
    col_widths = [10, 25, 18, 28, 12, 10, 14, 14, 40, 18, 14, 22, 28,
                  16, 16, 12, 12, 12, 12, 12, 12, 25, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    filepath = f"/tmp/horas_extras_{month:02d}_{year}.xlsx"
    wb.save(filepath)
    return filepath
 
async def start_csv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
 
    if not user_exists(user_id):
        await update.message.reply_text(
            "No estás registrado.\nUsá /registrar para crear tu cuenta."
        )
        return ConversationHandler.END
 
    await update.message.reply_text(
        "Ingresá el mes a consultar.\n"
        "Formato: MM/YYYY\n"
        "Ejemplo: 01/2026"
    )
    return REGISTER_MES
 
 
async def mes_csv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    texto = update.message.text.strip()
 
    try:
        fecha = datetime.strptime(texto, "%m/%Y")
    except ValueError:
        await update.message.reply_text(
            "Formato inválido.\nUsá MM/YYYY — Ejemplo: 03/2026"
        )
        return REGISTER_MES
 
    year = fecha.year
    month = fecha.month
    user_id = update.effective_user.id
 
    registros = get_overtime_by_moth(user_id, year, month)
 
    if not registros:
        await update.message.reply_text(
            "No se registraron horas extras en ese mes."
        )
        return ConversationHandler.END
 
    # Traer datos del usuario para completar la planilla
    user_data = get_user_by_telegram_id(user_id)
    if not user_data:
        await update.message.reply_text(
            "No se encontraron tus datos de usuario. Contactá al administrador."
        )
        return ConversationHandler.END
 
    await update.message.reply_text(
        f"Generando planilla para {month:02d}/{year}... ⏳"
    )
 
    filepath = _build_xlsx(user_data, registros, month, year)
    filename = os.path.basename(filepath)
 
    with open(filepath, "rb") as file:
        await update.message.reply_document(
            document=file,
            filename=filename,
            caption=f"📊 Planilla de horas extras — {month:02d}/{year}"
        )
 
    os.remove(filepath)
    return ConversationHandler.END
 
 
csv_handler = ConversationHandler(
    entry_points=[CommandHandler("csv", start_csv)],
    states={
        REGISTER_MES: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, mes_csv)
        ],
    },
    fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
)